import xlrd
import openpyxl
from traceback import print_exc
from tkFileDialog import askdirectory
from tkMessageBox import showinfo
from tkMessageBox import showerror
import Tkinter
import re
import os

def main():
    #crear diccionario para los boms-descripciones
    bomdict = {}

    #seleccionar la carpet
    #folder = "D:/myScripts/varios/bomextraction/testfiles"
    root = Tkinter.Tk()
    root.withdraw()
    folder = askdirectory(title="BOM CODE AND DESCRIPTION EXCTRACTION\n"
                                "SELECT A FOLDER CONTAINING PLS OR BOQS")

    #loop por todos los archivos de excel
    for fold, sub, files in os.walk(folder):
        for f in files:
            #quick check if it is an excel file
            if not '.xls' in f and not '.XLS' in f:
                continue

            print "scanning file: " + unicode(f)

            fname = os.path.join(fold,f)

            #leer archivos de excel y buscar boms y descripciones
            #guardar los resultados en un set
            try:
                wb = xlrd.open_workbook(fname)
            except:
                continue
            bomdict = xlreader(wb, bomdict)

    showinfo(title="PROCESS COMPLETED", message="total boms found: " + str(len(bomdict)))

    #escribir los resuultados del set en un libro de excel
    if len(bomdict) == 0:
        return

    write_results(folder, bomdict)
    os.startfile(os.path.join(folder, 'results.xlsx'))


def xlreader(workbook, bomdict):
    for sheet in workbook.sheets():
        try:
            part_col, desc_col, header_row = column_finder(sheet)
        except IndexError:
            continue

        if part_col and desc_col and header_row:
            #start collecting bom and description

            for r in range(header_row+1, sheet.nrows):
                bom = unicode(sheet.cell_value(r, part_col))
                if not re.search(r'^\d', bom):
                    continue
                if bom == '' or bom is None:
                    continue
                bomdict[bom] = sheet.cell_value(r, desc_col)

    return bomdict

def column_finder(sheet):
    '''toma una hoja como argumento y devuelve tres enteros
    que representan columna parte, columna descripcion y
    fila encabezado'''

    pattern_part = re.compile(r'^part\s?number', flags = re.IGNORECASE)
    pattern_desc = re.compile(r'^description', flags=re.IGNORECASE)

    for r in range(sheet.nrows):
        part_col = None
        desc_col = None
        for c in range(sheet.ncols):
            #busca part number y description en la misma fila
            #print sheet.name, r,c
            if pattern_part.search(unicode(sheet.cell_value(r, c))):
                part_col = c
            if pattern_desc.search(unicode(sheet.cell_value(r, c))):
                desc_col = c
        
        if part_col and desc_col:
            return part_col, desc_col, r
    

    return None, None, None

def write_results(folder, bomdict):
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.cell(row=1, column=1).value = "Part Number"
    sh.cell(row=1, column=2).value = "Description"

    nrow = 2

    for k in bomdict.keys():
        sh.cell(row=nrow, column=1).value = k
        sh.cell(row=nrow, column=2).value = bomdict[k]
        nrow+=1

    wb.save(os.path.join(folder,'results.xlsx'))


if __name__ == "__main__":
    # wb = xlrd.open_workbook("D:/myScripts/varios/bomextraction/testfiles/CNT_LTE_-IPRAN_R2-SWP_000000168923201605250001_35600501_BOQ.xlsx")
    # sheet = wb.sheet_by_name("L3-ATN 910I")
    # print column_finder(sheet)
    main()

    # try:
    #     main()
    # except Exception as e:
    #     showerror(title="AN ERROR OCURRED", message=repr(e))
